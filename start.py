import sys
import os
from PyQt5 import uic  # Импортируем uic
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import get_column_letter

import csv


class MyWidget(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('start.ui', self)  # Загружаем дизайн
        self.initUI()
        self.initMsgBox()
        self.loadDefaultSettings()
        self.addLog('Программа готова к работе')

    def initUI(self):
        self.btnFindData.setEnabled(False)
        self.btnRun.setEnabled(False)
        self.btnZ.hide()

        self.btnOpenFile.clicked.connect(self.openFile)
        self.btnFindData.clicked.connect(self.detectColumns)
        self.btnRun.clicked.connect(self.Run)

        self.deafult_styleSheet = self.styleSheet()
        self.work_styleSheet = "background-color: #FFC0CB"
        self.good_styleSheet = "background-color: #98FB98"

    def addLog(self, txt):
        self.txtLog.append(f"<strong>{self.log_n}:\t</strong>{txt}")
        # self.setStyleSheet(self.work_styleSheet)
        self.log_n += 1

    def initMsgBox(self):
        self.msg = QMessageBox(self)
        self.msg.setWindowTitle("Ошибка")
        self.msg.setIcon(QMessageBox.Warning)

    def loadDefaultSettings(self):
        self.wb = None
        self.sheet_plan_name = 'План'
        self.sheet_practice_name = 'Практики'
        self.row_start = 6
        self.col_name = 'C'  # Наименование
        self.col_exam = 'E'  # Экзамен
        self.col_zach = 'F'  # Зачет
        self.col_pract = 'G'  # Зачет с оц.
        self.col_proj = 'H'  # КП
        self.col_kurs = 'I'  # КР
        self.col_zet = 'J'  # Экспертное
        self.col_contact = 'O'  # Конт. раб.
        self.col_kafedra = 'CH'  # Наименование кафедры

        self.number_rows = None # количество строк  с данными
        self.number_columns = None # количество столбцов с данными

        self.log_n = 1

    def refreshData(self):
        self.sheet_plan_name =  self.txtPlan.text()
        self.row_start = int(self.txtRowStart.text())
        self.col_name = self.txtName.text()  # Наименование
        self.col_exam = self.txtExam.text()  # Экзамен
        self.col_zach = self.txtZach.text()  # Зачет
        self.col_pract = self.txtPract.text() # Зачет с оц.
        self.col_proj = self.txtProj.text() # КП
        self.col_kurs = self.txtKurs.text()  # КР
        self.col_zet = self.txtZet.text() # Экспертное
        self.col_contact = self.txtContact.text()  # Конт. раб.
        # self.col_kafedra = self.txtPlan.text()  # Наименование кафедры
    def openFile(self):
        try:
            self.fname = QFileDialog.getOpenFileName(
                self, 'Выбрать файл учебного плана', '',
                'Таблица в формате Excel (*.xlsx);;Все файлы (*)')[0]
            if self.fname:
                self.addLog(f'Пробуем открыть файл <span style="color:#0000ff">{self.fname}</span>')
                self.txtLog.repaint()
                self.workdir = os.path.dirname(self.fname)
                self.wb = load_workbook(self.fname)
                # Активируем нужный лист
                self.sheet = self.wb[self.sheet_plan_name]
                self.number_rows = self.sheet.max_row
                self.number_columns = self.sheet.max_column
                self.addLog(f'Файл открыт <span style="color:#0000ff">{self.fname}</span>')
                self.btnFindData.setEnabled(True)
                self.btnRun.setEnabled(True)
            else:
                self.addLog(f'<span style="color:#ff0000">пользователь забыл выбрать файл</span>')
        except:
            print('Ошибка в openFile')
            self.addLog(f'<span style="color:#ff0000">не получилось открыть файл</span>')
            self.msg.setText('Ошибка при открытие файла.')
            buttonAceptar = self.msg.addButton("Ok", QMessageBox.AcceptRole)
            self.msg.exec_()

    def detectColumns(self):
        print('Определяем столбцы')
        self.refreshData()
        self.col_kafedra = self.getColumn('Код')
        if self.col_kafedra:
            print(f'Нашли кафедру в столбце {self.col_kafedra}')
            self.addLog(f'Нашли кафедру в столбце {self.col_kafedra}')
            self.txtKafedra.setText(self.col_kafedra)


    def getColumn(self,txt):
        row=3
        for col in range(1,self.number_columns):
            cell=self.sheet[get_column_letter(col)+str(row)]
            if cell.value == txt:
                return get_column_letter(col+1)

    def closeEvent(self, event):
        if self.wb:
            print('Закрываем книгу')
            self.wb.close()
        else:
            print('Просто выходим')

    def Run(self):
        print('Запускаем обработку')
        self.addLog('Запускаем обработку')
        myData = [["№","Дисциплина", "Кол-во часов", "доля ставки"]]
        counter = 0
        for row in range(self.row_start,self.number_rows+1):
            cell_name = self.sheet[f"{self.col_name}{row}"].value
            cell_kafedra = self.sheet[f"{self.col_kafedra}{row}"].value
            col_pract =  self.sheet[f"{self.col_pract}{row}"].value
            col_zet =  self.sheet[f"{self.col_zet}{row}"].value
            col_exam =  self.sheet[f"{self.col_exam}{row}"].value
            col_zach =  self.sheet[f"{self.col_zach}{row}"].value
            col_proj =  self.sheet[f"{self.col_proj}{row}"].value
            col_kurs =  self.sheet[f"{self.col_kurs}{row}"].value
            col_contact =  self.sheet[f"{self.col_contact}{row}"].value
            # определяем, что это дисциплина а не модуль
            if (cell_name != None and cell_kafedra !=None):
                print(counter,cell_name)
                counter+=1
                tmp = []
                tmp.append(str(counter))
                tmp.append(cell_name)
                s = 0
                if 'квалификационной работы'.lower() in cell_name.lower():  # Если защита ВКР
                    s += 1
                if col_pract:  # Если практика
                    col_zet = self.sheet[f"{self.col_zet}{row}"].value
                    weeks = int(col_zet) / 1.5
                    s += 0.75 * weeks
                else:  # если просто дисциплина
                    # tmp.append(sheet[f"{col_contact}{row}"].value)
                    if col_exam:
                        s += 0.3 * len(col_exam)
                    if col_zach:
                        s += 0.2 * len(col_zach)
                    if col_proj:
                        s += 3
                    if col_kurs:
                        s += 2
                    if col_contact != None:
                        s += int(col_contact)
                s2 = round(s / 850, 2)
                tmp.append(str(s))
                tmp.append(str(s2))

                self.addLog(f'Обрабатываем дисциплину: '+" ".join(tmp))
                myData.append(tmp)

        myFile = open(self.fname+'.csv', 'w')
        with myFile:
            writer = csv.writer(myFile)
            writer.writerows(myData)
        myFile.close()
        self.addLog('<span style="color:#00ff00">Обработка завершена</span>')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyWidget()
    ex.show()
    sys.exit(app.exec_())
