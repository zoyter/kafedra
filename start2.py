import sys
import os
from PyQt5 import uic  # Импортируем uic
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import get_column_letter

import csv

'''
    Сначала берём из файла типа 09.03.01 Информатика и вычислительная техника_Технологии разработки мобильных приложений_ОФО_2021
    первую таблицу без заголовка, т.е. начина с данных
    вставляем в табличный процессор и сохраняем в файл qqq.xlsx
    после этого запускаем и в конце вводим кол-во людей во второй таблице
'''

def get_prepods(sheet):
    # Общее количество педагогических работников, реализующих основную образовательную программу, ________ чел.
    teachers = set()
    for row in range(1,sheet.max_row + 1):
        cell_name = sheet[f"C{row}"].value
        if cell_name:
            cell_name = cell_name.replace('\n',' ')
            cell_name = cell_name.replace('\r',' ')
            cell_name = cell_name.replace('  ','')
            cell_name = cell_name.lower()
            teachers.add(cell_name)
    return teachers

def get_stavki(sheet):
    # Общее количество ставок, занимаемых лицами, участвующими в реализации образовательной программы ______ ст.
    s = 0
    for row in range(1,sheet.max_row + 1):
        cell = sheet[f"i{row}"].value
        # print(row,type(cell))
        if cell:
            s+=cell
    print(s,'-------------------')
    return s

def get_scientist(sheet):
    # Доля педагогических работников (включая лиц, привлекаемых на иных условиях), имеющих ученую степень и/или ученое звание, составляет _______%;
    r = set()
    stavka = 0
    for row in range(1,sheet.max_row + 1):
        cell = sheet[f"E{row}"].value

        if cell:
            cell = cell.replace('\n',' ')
            cell = cell.replace('\r',' ')
            cell = cell.replace('  ','')
            cell = cell.lower()
            if 'кандидат' in cell or 'доктор' in cell or 'канд.' in cell:
                print(row)
                stavka += sheet[f"I{row}"].value
                cell_name = sheet[f"C{row}"].value
                cell_name = cell_name.replace('\n', ' ')
                cell_name = cell_name.replace('\r', ' ')
                cell_name = cell_name.replace('  ', '')
                cell_name = cell_name.lower()
                r.add(cell_name)
    # print(stavka)
    return stavka

def get_ext_rabotnik(sheet, sheet2):
    r = 0
    rabotodatel = []
    for row in range(1,sheet2.max_row + 1):
        name = sheet2[f"B{row}"].value
        if name:
            name = name.lower()  # получаем фамилию работодателя со второго листа
            if name not in rabotodatel:
                rabotodatel.append(name)
    print(rabotodatel)
    # Перебираем первый лист и ищем работодателей
    for row in range(1,sheet.max_row + 1):
        cell = sheet[f"C{row}"].value # имя препода с первого листа
        if cell:
            name2 = cell.lower().replace('\n',' ').replace('\t',' ').replace('  ',' ') #
            if name2 in rabotodatel:
                stavka = sheet[f"I{row}"].value
                if stavka:
                    r+=stavka
                    # print(name2, stavka,sep='\t')
    return r

fname = 'qqq.xlsx'
wb = load_workbook(fname)
sheet = wb['Лист1']
sheet2 = wb['Лист2']
number_rows = sheet.max_row
number_columns = sheet.max_column
print(number_rows,number_columns)
t=len(get_prepods(sheet))
s=get_stavki(sheet)

si = get_scientist(sheet)
wb.close()

print('Общее количество педагогических работников, реализующих основную образовательную программу, чел.')
print(f'\t {t}')
print('Общее количество ставок, занимаемых лицами, участвующими в реализации образовательной программы ______ ст.')
print(f'\t {s:.2f}')

print('Доля педагогических работников (включая лиц, привлекаемых на иных условиях), имеющих ученую степень и/или ученое звание, составляет _______%;')
print(f'\t{si} человек, {(si/s)*100:.2f} % ')


print('Доля работников из числа руководителей и (или) работников профильных организаций составляет _______%.')
rabot = get_ext_rabotnik(sheet,sheet2)
print(f'\t{rabot} человек, {(rabot/s)*100:.2f} % ')
